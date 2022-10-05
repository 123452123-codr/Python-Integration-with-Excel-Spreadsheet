from random import randint, random
from openpyxl import Workbook, load_workbook

wb = load_workbook('database.xlsx')
ws = wb.active

a = input("Enter a string to input into the database : ")

ws["A1"].value = a

print("String successfully inserted. Returned with code {}".format(randint(2312, 38217)))

wb.save('database.xlsx')
